'Functions for serializing and deserializing objects to csv and excel files'
from . import shininglive, account
from openpyxl import load_workbook


skillTypes = {}

for skill in shininglive.SkillType:
  for name in skill.value[1]:
    skillTypes[name.upper()] = skill



def load(filename):
  return load_workbook(filename, read_only = True);

def read_excel_cards(wb):
  cardsheet = wb['My Cards']
  cardnames = cardsheet['A3:A102']
  cards = {}
  for namecell in cardnames:
    name = namecell[0].value
    if name != '' and name != None:
      row = cardsheet[namecell[0].row]
      card = shininglive.Photo()
      card.name = name
      card.rarity = shininglive.Rarity[row[1].value.upper()]
      card.boy = shininglive.Boy[row[2].value.upper()]
      card.color = shininglive.Color[row[3].value.upper()]
      card.attributes[shininglive.Type.DANCE] = row[4].value
      card.attributes[shininglive.Type.VOCAL] = row[5].value
      card.attributes[shininglive.Type.CHARM] = row[6].value
      card.type = shininglive.Type[row[7].value.split(" ")[1].upper()]
      card.skill = shininglive.Skill(skillTypes[row[8].value.upper()], row[9].value)
      cards[name] = card
      #print(card)
     # card.rarity = shininglive.Rarity[
  #print(skillTypes)
  #print(cards)
  return cards

def read_excel_songs(wb, sheetname):
  songsheet = wb['Song Database']
  print(songsheet.max_row)
  for i in range(3, songsheet.max_row):
    songrow = songsheet[i]
    # need to replace symbols here
    songname = songrow[0].value
    if songname == '' or songname == None:
      if songrow[0].row > 50:
        break
      continue
    songname = songname.encode('cp850','replace').decode('cp850')
    song = shininglive.Song()
    song.name = songname
    song.notes = {}
    song.notes['hard'] = songrow[1].value
    song.notes['pro'] = songrow[2].value
    song.color = shininglive.Color[songrow[3].value.upper()]
    for j in range(4,15):
      singer = songrow[j].value
      if singer != None and singer != '':
        song.singers.add(shininglive.Boy[singer.upper()])
    print(song.notes)
    shininglive.SONG_LIST[song.name] = song
  return

def read_excel_event(wb, sheetname):
  profile = account.Profile()
  settings = wb['Settings']

  return

def read_excel_profile(wb, sheetname):
  return
